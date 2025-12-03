"""
===========================================================
 UAN AUTOMATOR v5.1 – Expedientes Oficiales con PDF Global
 Autor: Carlos Mosqueira (Arquitectos de Ideas 🤝 AI SOFTWARES)
 Descripción:
   - Automatiza la captura de alumnos para trámites UAN.
   - Llena 2 formatos de Excel (Machote y Formato de Gestores).
   - Crea una carpeta por alumno con sus documentos digitalizados.
   - Convierte imágenes/PDF a un único PDF COMPLETO_OFICIAL por alumno.
   - Genera un PDF GLOBAL con todos los expedientes oficiales.
===========================================================
"""

import shutil
import os
import sys
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Dict, List
import datetime

# =========================================================
#  IMPORTACIÓN DE LIBRERÍAS EXTERNAS
# =========================================================
try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from PIL import Image

    # Intentamos pypdf (nuevo), si no existe, usamos PyPDF2 (legacy)
    try:
        from pypdf import PdfWriter, PdfReader
    except ImportError:
        from PyPDF2 import PdfWriter, PdfReader

    from reportlab.lib.pagesizes import LETTER
    from reportlab.pdfgen import canvas

except ImportError as e:
    print(f"❌ FALTA INSTALAR LIBRERÍAS: {e}")
    print("Ejecuta en Terminal: pip3 install openpyxl pillow pypdf reportlab")
    sys.exit()

import tkinter as tk
from tkinter import filedialog

# =========================================================
#  CONFIGURACIÓN DE COLORES PARA CONSOLA (MAC / LINUX)
# =========================================================
class Color:
    VERDE = '\033[92m'
    AMARILLO = '\033[93m'
    ROJO = '\033[91m'
    AZUL = '\033[94m'
    RESET = '\033[0m'
    BOLD = '\033[1m'


# =========================================================
#  RUTAS Y ARCHIVOS PRINCIPALES
# =========================================================
TEMPLATE_MACHOTE = Path("MACHOTE DE TRAMITES.xlsx")
TEMPLATE_GESTORES = Path("FORMATO PARA LOS GESTORES.xlsx")

OUTPUT_MACHOTE = Path("MACHOTE_DE_TRAMITES_AUTOMATICO.xlsx")
OUTPUT_GESTORES = Path("FORMATO_PARA_LOS_GESTORES_AUTOMATICO.xlsx")

BASE_ALUMNOS_DIR = Path("ALUMNOS_UAN")
PDF_GLOBAL_NOMBRE = "PDF_GLOBAL_TODOS_LOS_ALUMNOS.pdf"


# =========================================================
#  DATA CLASSES – MODELOS DE DOCUMENTO Y ALUMNO
# =========================================================
@dataclass
class Documento:
    """
    Representa un documento entregado por el alumno
    (Acta, Certificado, CURP, Fotos, etc.)
    """
    nombre_clave: str              # Clave lógica interna, ej: "acta"
    nombre_archivo: str            # Base del nombre físico, sin extensión
    ruta_origen: Optional[Path] = None
    presentado: bool = False
    ruta_final_pdf: Optional[Path] = None


@dataclass
class Alumno:
    """
    Representa a un alumno y toda la información necesaria
    para trámites y formatos.
    """
    # Datos personales y académicos
    carrera: str
    curp: str
    nombre: str
    primer_apellido: str
    segundo_apellido: str
    institucion: str
    fecha_terminacion: str
    ciclo_escolar: str
    promedio: str

    # Trámites solicitados
    solicita_certificado: bool
    solicita_autenticacion: bool

    # Documentos entregados
    docs: Dict[str, Documento] = field(default_factory=dict)

    @property
    def nombre_completo(self) -> str:
        """Devuelve el nombre completo con apellidos."""
        return f"{self.nombre} {self.primer_apellido} {self.segundo_apellido}".strip()

    def get_excel_mark(self, key: str) -> str:
        """
        Devuelve "X" si el documento con clave `key` fue presentado.
        Se usa para llenar las columnas del Excel.
        """
        try:
            return "X" if self.docs[key].presentado else ""
        except KeyError:
            return ""


# =========================================================
#  UTILIDADES DE SISTEMA, ENTRADA Y GUI
# =========================================================
def limpiar_pantalla() -> None:
    """Limpia la pantalla de la terminal."""
    os.system('clear')


def validar_archivos_disponibles() -> None:
    """
    Verifica que existan las plantillas de Excel
    y que los archivos de salida no estén abiertos/bloqueados.
    """
    if not TEMPLATE_MACHOTE.exists():
        print(f"{Color.ROJO}❌ Error: No encuentro la plantilla '{TEMPLATE_MACHOTE}'{Color.RESET}")
        sys.exit()

    if not TEMPLATE_GESTORES.exists():
        print(f"{Color.ROJO}❌ Error: No encuentro la plantilla '{TEMPLATE_GESTORES}'{Color.RESET}")
        sys.exit()

    for f in [OUTPUT_MACHOTE, OUTPUT_GESTORES]:
        if f.exists():
            try:
                with open(f, "r+b"):
                    pass
            except OSError:
                print(f"{Color.ROJO}❌ El archivo '{f.name}' está abierto. Ciérralo y vuelve a correr el programa.{Color.RESET}")
                sys.exit()


def seleccionar_archivo(titulo: str) -> Optional[Path]:
    """
    Abre una ventana nativa para seleccionar un archivo.
    Se forza a estar al frente (topmost) para que en Mac no se esconda.
    """
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    ruta = filedialog.askopenfilename(title=f"📂 SELECCIONA: {titulo}")
    root.destroy()

    if ruta:
        return Path(ruta)
    return None


def input_clean(msg: str) -> str:
    """
    Input con formato:
    - Mensaje en negritas.
    - Se regresa en mayúsculas y sin espacios extras.
    """
    return input(f"{Color.BOLD}{msg}{Color.RESET} ").strip().upper()


def input_si_no(pregunta: str) -> bool:
    """
    Pregunta cerrada tipo S/N.
    Regresa True si la respuesta comienza con 'S',
    False si comienza con 'N'.
    """
    while True:
        resp = input(f"{pregunta} [S/N]: ").strip().upper()
        if resp.startswith("S"):
            return True
        if resp.startswith("N"):
            return False


def safe_copy(src: Path, dest: Path) -> None:
    """
    Copia segura de archivos:
    - Ignora si src no existe.
    - Evita copiar un archivo sobre sí mismo.
    """
    try:
        if not src or not src.exists():
            return
        if src.resolve() == dest.resolve():
            return
        shutil.copy(src, dest)
    except Exception as e:
        print(f"{Color.ROJO}⚠️ Error copiando {src.name}: {e}{Color.RESET}")


# =========================================================
#  MOTOR DE PDF – CONVERSIÓN Y UNIÓN
# =========================================================
def convertir_imagen_a_pdf(imagen_path: Path, pdf_path: Path) -> bool:
    """
    Convierte una imagen (PNG/JPG/etc.) a PDF.
    Regresa True si la conversión fue exitosa.
    """
    try:
        image = Image.open(imagen_path)
        if image.mode in ("RGBA", "P"):
            image = image.convert("RGB")
        image.save(pdf_path, "PDF", resolution=100.0)
        return True
    except Exception as e:
        print(f"{Color.AMARILLO}   ⚠️ Error convirtiendo imagen {imagen_path.name}: {e}{Color.RESET}")
        return False


def procesar_documento_a_pdf(doc: Documento, carpeta_destino: Path) -> Optional[Path]:
    """
    Copia el documento original a la carpeta del alumno
    y garantiza un PDF asociado (si es posible).

    Lógica:
      - Si es PDF: se copia directamente como PDF final.
      - Si es imagen: se convierte a PDF.
      - Otros formatos: se notifican como no soportados.
    """
    if not doc.presentado or not doc.ruta_origen:
        return None

    ext = doc.ruta_origen.suffix.lower()
    nombre_base = doc.nombre_archivo  # ya incluye CURP en el flujo principal

    # Copia de evidencia con extensión original
    archivo_final = carpeta_destino / f"{nombre_base}{ext}"
    safe_copy(doc.ruta_origen, archivo_final)

    # Ruta destino del PDF "limpio"
    pdf_final = carpeta_destino / f"{nombre_base}.pdf"

    if ext == ".pdf":
        safe_copy(doc.ruta_origen, pdf_final)
        return pdf_final

    if ext in (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"):
        if convertir_imagen_a_pdf(doc.ruta_origen, pdf_final):
            return pdf_final

    print(f"{Color.AMARILLO}   ⚠️ Formato no soportado para PDF automático: {doc.ruta_origen.name}{Color.RESET}")
    return None


def unir_pdfs_en_uno(lista_pdfs: List[Path], salida: Path) -> None:
    """
    Une una lista de PDFs en un solo archivo llamado `salida`.
    El orden en la lista define el orden dentro del PDF final.
    """
    merger = PdfWriter()
    for pdf in lista_pdfs:
        try:
            merger.append(str(pdf))
        except Exception as e:
            print(f"{Color.AMARILLO}   ⚠️ No se pudo unir el PDF {pdf.name}: {e}{Color.RESET}")
    with open(salida, "wb") as f:
        merger.write(f)


def generar_machote_pdf_alumno(al: Alumno, carpeta: Path, gestor: str) -> Path:
    """
    Genera un PDF "tipo machote" con los datos clave del alumno.
    Este PDF será siempre la primera página del expediente oficial.
    """
    pdf_path = carpeta / f"{al.curp}_MACHOTE_TRAMITES.pdf"
    c = canvas.Canvas(str(pdf_path), pagesize=LETTER)
    width, height = LETTER

    x = 60
    y = height - 60

    # Encabezado institucional
    c.setFont("Helvetica-Bold", 14)
    c.drawString(x, y, "UNIVERSIDAD DE AMÉRICA DEL NORTE")
    y -= 22
    c.drawString(x, y, "Formato de Solicitud para Elaboración de Certificados")
    y -= 30

    # Datos generales
    c.setFont("Helvetica", 10)
    hoy = datetime.date.today().strftime("%d/%m/%Y")
    c.drawString(x, y, f"Fecha de captura: {hoy}")
    y -= 18
    c.drawString(x, y, f"Gestor: {gestor}")
    y -= 25

    def linea(txt: str):
        """Imprime una línea con salto automático de página."""
        nonlocal y
        if y < 60:
            c.showPage()
            y = height - 60
            c.setFont("Helvetica", 10)
        c.drawString(x, y, txt)
        y -= 14

    # Bloque datos del alumno
    c.setFont("Helvetica-Bold", 11)
    linea("DATOS DEL ALUMNO")
    c.setFont("Helvetica", 10)
    linea(f"Nombre completo: {al.nombre_completo}")
    linea(f"CURP: {al.curp}")
    linea(f"Carrera: {al.carrera}")
    linea(f"Institución de procedencia: {al.institucion}")
    linea(f"Fecha de terminación de estudios: {al.fecha_terminacion}")
    linea(f"Ciclo escolar: {al.ciclo_escolar}")
    linea(f"Promedio: {al.promedio}")

    # Bloque trámites solicitados
    y -= 6
    c.setFont("Helvetica-Bold", 11)
    linea("TRÁMITES SOLICITADOS")
    c.setFont("Helvetica", 10)
    linea(f"• Certificado: {'X' if al.solicita_certificado else ''}")
    linea(f"• Autenticación: {'X' if al.solicita_autenticacion else ''}")

    # Bloque documentos entregados
    y -= 6
    c.setFont("Helvetica-Bold", 11)
    linea("DOCUMENTOS ENTREGADOS")
    c.setFont("Helvetica", 10)
    linea(f"- Certificado (Sec/Bach): {al.get_excel_mark('cert_sec')}")
    linea(f"- Acta de nacimiento: {al.get_excel_mark('acta')}")
    linea(f"- CURP impresa: {al.get_excel_mark('curp_doc')}")
    linea(f"- Fotos tamaño infantil: {al.get_excel_mark('fotos_inf')}")
    linea(f"- Fotos tamaño credencial: {al.get_excel_mark('fotos_cred')}")
    linea(f"- Fotos tamaño título: {al.get_excel_mark('fotos_tit')}")
    linea(f"- INE: {al.get_excel_mark('ine')}")
    linea(f"- Comprobante de domicilio: {al.get_excel_mark('comp_dom')}")
    linea(f"- INE tutor: {al.get_excel_mark('ine_tutor')}")

    # Firma del Gestor
    y -= 20
    linea("Nombre y firma del Gestor: _______________________________")

    c.showPage()
    c.save()
    return pdf_path


# =========================================================
#  MOTOR DE EXCEL – CARGA, LIMPIEZA Y ESCRITURA
# =========================================================
class GestorExcel:
    """
    Encapsula toda la lógica de interacción con los libros de Excel:
    - Machote de trámites.
    - Formato para gestores.
    """

    def __init__(self, gestor: str):
        self.gestor = gestor
        print(f"{Color.AZUL}⏳ Cargando plantillas Excel...{Color.RESET}")

        # --- LIBRO: MACHOTE ---
        if OUTPUT_MACHOTE.exists():
            self.wb_machote = load_workbook(OUTPUT_MACHOTE)
        else:
            self.wb_machote = load_workbook(TEMPLATE_MACHOTE)
            self._limpiar_plantilla(self.wb_machote["formato corregido"], "Ejemplo")

        self.ws_machote = self.wb_machote["formato corregido"]
        self.ws_machote["B29"] = self.gestor  # Nombre del gestor impreso

        # --- LIBRO: GESTORES ---
        if OUTPUT_GESTORES.exists():
            self.wb_gestores = load_workbook(OUTPUT_GESTORES)
        else:
            self.wb_gestores = load_workbook(TEMPLATE_GESTORES)
            self._limpiar_plantilla(self.wb_gestores["TITULOS"], "ES MUY IMPORTANTE")

        self.ws_gestores = self.wb_gestores["TITULOS"]
        self._firmar_gestores()

        # Punteros a la siguiente fila libre
        self.row_m = self._find_next_row(self.ws_machote, col_check=1, start=10)
        self.row_g = self._find_next_row(self.ws_gestores, col_check=2, start=3)

    def _limpiar_plantilla(self, ws, keyword: str) -> None:
        """
        Limpia registros de ejemplo sin romper celdas combinadas.
        Usa la presencia de una palabra clave como referencia.
        """
        start_clean = None
        for r in range(1, 80):
            val = ws.cell(row=r, column=1).value
            if val and isinstance(val, str) and keyword in val:
                start_clean = r
                break

        if start_clean is None:
            return

        if keyword == "Ejemplo":  # Machote
            rango = range(start_clean + 1, start_clean + 20)
            max_col = 18
        else:  # Gestores
            rango = range(3, start_clean)
            max_col = 16

        for r in rango:
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None

    def _find_next_row(self, ws, col_check: int, start: int) -> int:
        """
        Encuentra la siguiente fila vacía revisando una columna específica.
        """
        r = start
        while ws.cell(row=r, column=col_check).value:
            r += 1
        return r

    def _firmar_gestores(self) -> None:
        """
        Reemplaza el texto 'persona que solicita' por 'Gestor: {nombre}'.
        """
        for r in range(1, 60):
            for c in range(1, 15):
                v = self.ws_gestores.cell(row=r, column=c).value
                if v and isinstance(v, str) and "persona que solicita" in v.lower():
                    self.ws_gestores.cell(row=r, column=c).value = f"Gestor: {self.gestor}"
                    return

    def agregar_alumno(self, al: Alumno) -> None:
        """
        Inserta los datos del alumno tanto en el Machote como en el
        Formato para Gestores.
        """
        # --- Machote ---
        ws = self.ws_machote
        r = self.row_m

        prev = ws.cell(row=r - 1, column=1).value
        num = int(prev) + 1 if isinstance(prev, (int, float)) else 1

        row_data = [
            num,
            al.curp,
            al.nombre_completo,
            al.ciclo_escolar,
            al.fecha_terminacion,
            al.get_excel_mark("acta"),
            al.get_excel_mark("cert_sec"),
            al.get_excel_mark("curp_doc"),
            al.get_excel_mark("fotos_inf"),
            al.get_excel_mark("fotos_cred"),
            al.get_excel_mark("fotos_tit"),
            al.get_excel_mark("ine"),
            al.get_excel_mark("comp_dom"),
            al.carrera,
            "X" if al.solicita_certificado else "",
            "X" if al.solicita_autenticacion else "",
            al.promedio,
        ]

        for i, val in enumerate(row_data, 1):
            ws.cell(row=r, column=i).value = val
        self.row_m += 1

        # --- Formato Gestores ---
        ws = self.ws_gestores
        r = self.row_g

        any_fotos = "X" if (
            al.docs["fotos_inf"].presentado
            or al.docs["fotos_cred"].presentado
            or al.docs["fotos_tit"].presentado
        ) else ""

        row_data_g = [
            al.carrera,
            al.curp,
            al.nombre,
            al.primer_apellido,
            al.segundo_apellido,
            al.institucion,
            al.fecha_terminacion,
            al.get_excel_mark("cert_sec"),
            al.get_excel_mark("acta"),
            any_fotos,
            al.get_excel_mark("ine"),
            al.get_excel_mark("comp_dom"),
            al.get_excel_mark("ine_tutor"),
            al.ciclo_escolar,
            al.promedio,
        ]

        for i, val in enumerate(row_data_g, 1):
            ws.cell(row=r, column=i).value = val
        self.row_g += 1

    def guardar_todo(self) -> None:
        """Guarda los dos libros de Excel en disco."""
        print(f"\n{Color.AZUL}💾 Guardando libros de Excel...{Color.RESET}")
        self.wb_machote.save(OUTPUT_MACHOTE)
        self.wb_gestores.save(OUTPUT_GESTORES)
        print(f"{Color.VERDE}✅ Excels actualizados.{Color.RESET}")


# =========================================================
#  FUNCIÓN PRINCIPAL – FLUJO COMPLETO
# =========================================================
def main() -> None:
    """
    Punto de entrada del programa.
    Maneja el ciclo de captura de alumnos y el cierre global (PDF maestro).
    """
    limpiar_pantalla()
    validar_archivos_disponibles()

    print(f"{Color.VERDE}=============================================")
    print(f"   🚀 UAN AUTOMATOR v5.1 (OFICIAL + MACHOTE)")
    print(f"============================================={Color.RESET}")

    gestor = input_clean("👤 Nombre del Gestor:")
    excel_engine = GestorExcel(gestor)

    # Diccionario de configuración de documentos:
    # clave interna  -> (nombre base de archivo, texto de pregunta)
    conf_docs: Dict[str, tuple[str, str]] = {
        "cert_sec": ("CERTIFICADO_BACH", "¿Entregó Certificado (Sec/Bach)?"),
        "acta": ("ACTA_NACIMIENTO", "¿Entregó Acta de Nacimiento?"),
        "curp_doc": ("CURP", "¿Entregó CURP Impresa?"),
        "fotos_inf": ("FOTOS_INFANTIL", "¿Entregó Fotos Infantiles?"),
        "fotos_cred": ("FOTOS_CREDENCIAL", "¿Entregó Fotos Credencial?"),
        "fotos_tit": ("FOTOS_TITULO", "¿Entregó Fotos Título?"),
        "ine": ("INE", "¿Entregó INE?"),
        "comp_dom": ("COMP_DOMICILIO", "¿Entregó Comp. Domicilio?"),
        "ine_tutor": ("INE_TUTOR", "¿Entregó INE Tutor?"),
    }

    # Lista de PDF oficiales generados por alumno
    todos_los_pdfs_generados: List[Path] = []

    # ----------------------------
    # LOOP PRINCIPAL – ALUMNO
    # ----------------------------
    while True:
        limpiar_pantalla()
        print(f"{Color.AZUL}🎓 NUEVO ALUMNO (Gestor: {gestor}){Color.RESET}\n")

        # 1. Captura de datos personales
        curp = input_clean("📌 CURP:")
        print(f"   ... Procesando CURP: {curp}")

        al = Alumno(
            curp=curp,
            carrera=input_clean("📌 Carrera:"),
            nombre=input_clean("📌 Nombre(s):"),
            primer_apellido=input_clean("📌 Primer Apellido:"),
            segundo_apellido=input_clean("📌 Segundo Apellido:"),
            institucion=input_clean("📌 Institución Procedencia:"),
            fecha_terminacion=input_clean("📅 Fecha Terminación (dd/mm/aaaa):"),
            ciclo_escolar=input_clean("📅 Ciclo Escolar:"),
            promedio=input_clean("📈 Promedio:"),
            solicita_certificado=input_si_no("\n¿Solicita CERTIFICADO?"),
            solicita_autenticacion=input_si_no("¿Solicita AUTENTICACIÓN?"),
        )

        # 2. Captura de documentos
        print(f"\n{Color.AMARILLO}📂 RECEPCIÓN DE DOCUMENTOS{Color.RESET}")
        print("   (Si respondes 'S', se abrirá la ventana para elegir el archivo)")

        # Manejo especial de "fotografías completas"
        fotos_pack = input_si_no("¿Entregó FOTOGRAFÍAS COMPLETAS (un solo archivo con todas)?")
        handled_photos = set()

        if fotos_pack:
            ruta = seleccionar_archivo("FOTOGRAFÍAS (Pack Completo)")
            for k in ["fotos_inf", "fotos_cred", "fotos_tit"]:
                al.docs[k] = Documento(k, conf_docs[k][0], ruta, True)
                handled_photos.add(k)
        else:
            print("👉 No olvides mandar hacer tus fotografías al 56 38 38 9671 para que cumplan con los requisitos.\n")

        # Resto de documentos (y fotos individuales si no hubo pack)
        for key, (nom_archivo, pregunta) in conf_docs.items():
            if key in handled_photos:
                continue

            presentado = input_si_no(pregunta)
            ruta = None
            if presentado:
                ruta = seleccionar_archivo(nom_archivo)
                if ruta:
                    print(f"      ✅ {ruta.name}")
                else:
                    print(f"      ⚠️  Marcado como entregado sin archivo digital.")

            al.docs[key] = Documento(key, nom_archivo, ruta, presentado)

        # 3. Procesamiento de expediente del alumno
        print(f"\n{Color.AZUL}⚙️  Procesando Expediente...{Color.RESET}")

        # A) Crear carpeta específica del alumno
        BASE_ALUMNOS_DIR.mkdir(exist_ok=True)
        carpeta_al = BASE_ALUMNOS_DIR / f"{al.curp}_{al.nombre_completo.replace(' ', '_')}"
        carpeta_al.mkdir(exist_ok=True)

        # B) Generar PDF "machote" del alumno (primera hoja del expediente)
        pdfs_alumno: List[Path] = []
        pdf_machote = generar_machote_pdf_alumno(al, carpeta_al, gestor)
        pdfs_alumno.append(pdf_machote)

        # C) Procesar todos los documentos y convertirlos a PDF
        for key, doc in al.docs.items():
            # Prefijo CURP en el nombre físico de archivo
            doc.nombre_archivo = f"{al.curp}_{doc.nombre_archivo}"
            pdf_path = procesar_documento_a_pdf(doc, carpeta_al)
            if pdf_path and pdf_path not in pdfs_alumno:
                doc.ruta_final_pdf = pdf_path
                pdfs_alumno.append(pdf_path)

        # D) Crear PDF oficial único del alumno (machote + docs)
        if pdfs_alumno:
            pdf_completo_al = carpeta_al / f"{al.curp}_COMPLETO_OFICIAL.pdf"
            unir_pdfs_en_uno(pdfs_alumno, pdf_completo_al)
            print(f"   📄 PDF oficial del alumno: {pdf_completo_al.name}")
            todos_los_pdfs_generados.append(pdf_completo_al)
        else:
            print(f"{Color.AMARILLO}   ⚠️ No se generó PDF completo (sin documentos).{Color.RESET}")

        # E) Registrar alumno en ambos Excels
        excel_engine.agregar_alumno(al)

        print(f"{Color.VERDE}✅ Alumno procesado correctamente.{Color.RESET}")

        # ¿Otro alumno?
        if not input_si_no(f"\n{Color.AMARILLO}¿Capturar otro alumno?{Color.RESET}"):
            break

    # =====================================================
    #  CIERRE GLOBAL – GUARDAR EXCEL Y PDF MAESTRO
    # =====================================================
    # 1. Guardar Excels llenos
    excel_engine.guardar_todo()

    # 2. Construir/Actualizar PDF GLOBAL con todos los alumnos
    if todos_los_pdfs_generados:
        print(f"\n{Color.AZUL}📚 Generando PDF GLOBAL de todos los alumnos...{Color.RESET}")
        ruta_global = BASE_ALUMNOS_DIR / PDF_GLOBAL_NOMBRE

        lista_final: List[Path] = []

        # Si existe un PDF global anterior, lo preservamos como .bak y lo incluimos primero
        if ruta_global.exists():
            print("   ℹ️  Anexando a PDF Global existente (respaldo .bak)...")
            bak = ruta_global.with_suffix(".bak.pdf")
            ruta_global.rename(bak)
            lista_final.append(bak)

        # Agregamos los expedientes oficiales de la sesión actual
        lista_final.extend(todos_los_pdfs_generados)

        unir_pdfs_en_uno(lista_final, ruta_global)
        print(f"{Color.VERDE}✅ PDF GLOBAL GENERADO: {ruta_global}{Color.RESET}")

    print(f"\n{Color.VERDE}✨✨ PROCESO TERMINADO CON ÉXITO ✨✨{Color.RESET}")
    input("Presiona Enter para salir.")


# =========================================================
#  EJECUCIÓN DIRECTA
# =========================================================
if __name__ == "__main__":
    main()

import os
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Dict, List
from io import BytesIO

import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from PIL import Image

# PDF engine (pypdf o PyPDF2)
try:
    from pypdf import PdfWriter
except ImportError:
    from PyPDF2 import PdfWriter


# ==============================
#   CONFIGURACIÓN BÁSICA
# ==============================

st.set_page_config(
    page_title="UAN Automator",
    page_icon="🎓",
    layout="wide",
)

TEMPLATE_MACHOTE = Path("MACHOTE DE TRAMITES.xlsx")
TEMPLATE_GESTORES = Path("FORMATO PARA LOS GESTORES.xlsx")

BASE_ALUMNOS_DIR = Path("ALUMNOS_UAN")
BASE_ALUMNOS_DIR.mkdir(exist_ok=True)


# ==============================
#   MODELOS DE DATOS
# ==============================

@dataclass
class Documento:
    clave: str
    nombre_archivo: str
    ruta_origen: Optional[Path] = None
    presentado: bool = False
    ruta_final_pdf: Optional[Path] = None


@dataclass
class Alumno:
    carrera: str
    curp: str
    nombre: str
    primer_apellido: str
    segundo_apellido: str
    institucion: str
    fecha_terminacion: str
    ciclo_escolar: str
    promedio: str
    solicita_certificado: bool
    solicita_autenticacion: bool
    docs: Dict[str, Documento] = field(default_factory=dict)

    @property
    def nombre_completo(self) -> str:
        return f"{self.nombre} {self.primer_apellido} {self.segundo_apellido}".strip()

    def mark(self, key: str) -> str:
        return "X" if self.docs.get(key) and self.docs[key].presentado else ""


# ==============================
#   UTILIDADES PDF / ARCHIVOS
# ==============================

def convertir_imagen_a_pdf(imagen_path: Path, pdf_path: Path) -> bool:
    try:
        im = Image.open(imagen_path)
        if im.mode in ("RGBA", "P"):
            im = im.convert("RGB")
        im.save(pdf_path, "PDF", resolution=100.0)
        return True
    except Exception as e:
        st.warning(f"⚠️ Error convirtiendo imagen {imagen_path.name}: {e}")
        return False


def procesar_documento_a_pdf(doc: Documento, carpeta_destino: Path) -> Optional[Path]:
    """Copia el archivo a la carpeta del alumno y genera un PDF."""
    if not doc.presentado or not doc.ruta_origen:
        return None

    ext = doc.ruta_origen.suffix.lower()
    destino_original = carpeta_destino / f"{doc.nombre_archivo}{ext}"
    destino_original.write_bytes(doc.ruta_origen.read_bytes())

    destino_pdf = carpeta_destino / f"{doc.nombre_archivo}.pdf"

    if ext == ".pdf":
        destino_pdf.write_bytes(doc.ruta_origen.read_bytes())
        return destino_pdf
    elif ext in (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"):
        if convertir_imagen_a_pdf(doc.ruta_origen, destino_pdf):
            return destino_pdf

    return None


def unir_pdfs_en_uno(lista_pdfs: List[Path], salida: Path) -> bool:
    merger = PdfWriter()
    for pdf in lista_pdfs:
        try:
            merger.append(str(pdf))
        except Exception as e:
            st.warning(f"⚠️ No se pudo unir {pdf.name}: {e}")
    if len(merger.pages) == 0:
        return False
    with open(salida, "wb") as f:
        merger.write(f)
    return True


# ==============================
#   MOTOR EXCEL
# ==============================

class GestorExcel:
    def __init__(self, gestor_nombre: str):
        self.gestor = gestor_nombre.upper()
        self._cargar_libros()

    def _cargar_libros(self):
        # Machote
        self.wb_machote = load_workbook(TEMPLATE_MACHOTE)
        ws = self.wb_machote["formato corregido"]
        self._limpiar_machote(ws)
        ws["B29"] = self.gestor
        self.ws_machote = ws

        # Gestores
        self.wb_gestores = load_workbook(TEMPLATE_GESTORES)
        ws2 = self.wb_gestores["TITULOS"]
        self._limpiar_gestores(ws2)
        self._firmar_gestores(ws2)
        self.ws_gestores = ws2

        self.row_m = self._find_next_row(self.ws_machote, 1, start=10)
        self.row_g = self._find_next_row(self.ws_gestores, 2, start=3)

    def _find_next_row(self, ws, col_check: int, start: int) -> int:
        r = start
        while ws.cell(row=r, column=col_check).value not in (None, ""):
            r += 1
        return r

    def _limpiar_machote(self, ws):
        start_row = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip().lower().startswith("ejemplo"):
                start_row = r
                break
        if start_row:
            for r in range(start_row + 1, start_row + 25):
                for c in range(1, 18):
                    cell = ws.cell(row=r, column=c)
                    if not isinstance(cell, MergedCell):
                        cell.value = None

    def _limpiar_gestores(self, ws):
        limit_row = 30
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and "ES MUY IMPORTANTE" in v.upper():
                limit_row = r
                break
        for r in range(3, limit_row):
            for c in range(1, 16):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None

    def _firmar_gestores(self, ws):
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and "persona que solicita" in v.lower():
                    ws.cell(row=r, column=c).value = f"Gestor: {self.gestor}"
                    return

    def agregar_alumno(self, al: Alumno):
        # Machote
        ws = self.ws_machote
        r = self.row_m
        prev = ws.cell(row=r - 1, column=1).value
        num = int(prev) + 1 if isinstance(prev, (int, float)) else 1

        datos_m = [
            num, al.curp, al.nombre_completo, al.ciclo_escolar, al.fecha_terminacion,
            al.mark("acta"), al.mark("cert_sec"), al.mark("curp_doc"),
            al.mark("fotos_inf"), al.mark("fotos_cred"), al.mark("fotos_tit"),
            al.mark("ine"), al.mark("comp_dom"), al.carrera,
            "X" if al.solicita_certificado else "",
            "X" if al.solicita_autenticacion else "",
            al.promedio
        ]
        for i, v in enumerate(datos_m, 1):
            ws.cell(row=r, column=i).value = v
        self.row_m += 1

        # Gestores
        ws2 = self.ws_gestores
        r2 = self.row_g
        any_fotos = "X" if (
            al.docs["fotos_inf"].presentado
            or al.docs["fotos_cred"].presentado
            or al.docs["fotos_tit"].presentado
        ) else ""

        datos_g = [
            al.carrera, al.curp, al.nombre, al.primer_apellido, al.segundo_apellido,
            al.institucion, al.fecha_terminacion, al.mark("cert_sec"), al.mark("acta"),
            any_fotos, al.mark("ine"), al.mark("comp_dom"), al.mark("ine_tutor"),
            al.ciclo_escolar, al.promedio
        ]
        for i, v in enumerate(datos_g, 1):
            ws2.cell(row=r2, column=i).value = v
        self.row_g += 1

    def get_excels_as_bytes(self):
        buf_m = BytesIO()
        buf_g = BytesIO()
        self.wb_machote.save(buf_m)
        self.wb_gestores.save(buf_g)
        buf_m.seek(0)
        buf_g.seek(0)
        return buf_m, buf_g


# ==============================
#   INTERFAZ STREAMLIT
# ==============================

st.title("🎓 UAN AUTOMATOR – Certificados y Expedientes")

st.markdown(
    """
**Automatizador oficial de trámites UAN**

- Genera carpeta de cada alumno  
- Convierte documentos a PDF y crea el **EXPEDIENTE COMPLETO**  
- Llena automáticamente el **MACHOTE DE TRÁMITES**  
- Llena el **FORMATO PARA LOS GESTORES**  
"""
)

st.sidebar.title("⚙️ Configuración")
gestor_nombre = st.sidebar.text_input("Nombre del Gestor", value=st.session_state.get("gestor", ""))
if gestor_nombre.strip():
    st.session_state["gestor"] = gestor_nombre.strip().upper()

st.sidebar.info("Completa los datos y presiona **Generar** para crear el expediente.")

if "engine" not in st.session_state and gestor_nombre.strip():
    st.session_state["engine"] = GestorExcel(gestor_nombre)

engine: Optional[GestorExcel] = st.session_state.get("engine")

if not gestor_nombre.strip():
    st.warning("👉 Escribe primero el **Nombre del Gestor** en la barra lateral.")
    st.stop()

st.markdown("---")
st.header("1️⃣ Datos del alumno")

col1, col2, col3 = st.columns(3)

with col1:
    curp = st.text_input("CURP").upper().strip()
    nombre = st.text_input("Nombre(s)").upper().strip()
    institucion = st.text_input("Institución de Procedencia").upper().strip()

with col2:
    p_ap = st.text_input("Primer Apellido").upper().strip()
    ciclo = st.text_input("Ciclo Escolar").upper().strip()
    fecha_terminacion = st.text_input("Fecha de Terminación (dd/mm/aaaa)").strip()

with col3:
    s_ap = st.text_input("Segundo Apellido").upper().strip()
    promedio = st.text_input("Promedio").strip()
    carrera = st.text_input("Carrera").upper().strip()

st.markdown("---")
st.header("2️⃣ Trámites solicitados")

colt1, colt2 = st.columns(2)
with colt1:
    solicita_cert = st.checkbox("Solicita CERTIFICADO", value=True)
with colt2:
    solicita_auth = st.checkbox("Solicita AUTENTICACIÓN", value=True)

st.markdown("---")
st.header("3️⃣ Documentos entregados")

st.info("Marca lo que entregó el alumno y sube el archivo correspondiente. Las fotos pueden subirse en un solo archivo o separados.")

conf_docs = {
    "cert_sec": ("CERTIFICADO_BACH", "Certificado (Sec/Bach)"),
    "acta": ("ACTA_NACIMIENTO", "Acta de nacimiento"),
    "curp_doc": ("CURP", "CURP Impresa"),
    "fotos_inf": ("FOTOS_INFANTIL", "Fotos tamaño infantil"),
    "fotos_cred": ("FOTOS_CREDENCIAL", "Fotos tamaño credencial"),
    "fotos_tit": ("FOTOS_TITULO", "Fotos tamaño título"),
    "ine": ("INE", "INE"),
    "comp_dom": ("COMP_DOMICILIO", "Comprobante de domicilio"),
    "ine_tutor": ("INE_TUTOR", "INE padre o tutor (si aplica)"),
}

cols = st.columns(3)
uploaded_info = {}

# Pregunta especial: fotos completas
fotos_completas = st.checkbox("✅ Entregó FOTOGRAFÍAS COMPLETAS en un solo archivo")
archivo_fotos_completo = None
if fotos_completas:
    archivo_fotos_completo = st.file_uploader(
        "Archivo único de fotografías (infantil, credencial, título)",
        type=["pdf", "png", "jpg", "jpeg"],
        key="fotos_pack",
    )

for i, (key, (nombre_archivo, label)) in enumerate(conf_docs.items()):
    col = cols[i % 3]
    with col:
        entregado = st.checkbox(label, key=f"chk_{key}")
        archivo = None

        # Si es una de las fotos y marcó pack completo
        if fotos_completas and key in ("fotos_inf", "fotos_cred", "fotos_tit"):
            entregado = bool(archivo_fotos_completo)
            archivo = archivo_fotos_completo
        else:
            if entregado:
                archivo = st.file_uploader(
                    f"Archivo: {label}",
                    type=["pdf", "png", "jpg", "jpeg"],
                    key=f"file_{key}",
                )
        uploaded_info[key] = (entregado, archivo)

if fotos_completas:
    st.caption("📸 Tip: Si no entregó fotos completas, desmarca la opción y súbelas por separado.")

st.markdown("---")
st.header("4️⃣ Generar expediente y formatos")

generar = st.button("🚀 GENERAR ARCHIVOS DEL ALUMNO", type="primary")

if generar:
    # Validaciones mínimas
    if not curp or not nombre or not p_ap:
        st.error("La CURP, Nombre y Primer Apellido son obligatorios.")
    else:
        # Crear objeto Alumno
        alumno = Alumno(
            carrera=carrera,
            curp=curp,
            nombre=nombre,
            primer_apellido=p_ap,
            segundo_apellido=s_ap,
            institucion=institucion,
            fecha_terminacion=fecha_terminacion,
            ciclo_escolar=ciclo,
            promedio=promedio,
            solicita_certificado=solicita_cert,
            solicita_autenticacion=solicita_auth,
        )

        # Carpeta del alumno
        carpeta_alumno = BASE_ALUMNOS_DIR / f"{alumno.curp}_{alumno.nombre_completo.replace(' ', '_')}"
        carpeta_alumno.mkdir(exist_ok=True)

        # Guardar archivos subidos temporalmente y convertir a Documento
        pdfs_alumno: List[Path] = []
        for key, (nombre_archivo, _) in conf_docs.items():
            entregado, archivo = uploaded_info[key]
            ruta_tmp = None

            if archivo is not None:
                # Guardar archivo subido en disco
                ext = Path(archivo.name).suffix.lower()
                ruta_tmp = carpeta_alumno / f"{alumno.curp}_{nombre_archivo}{ext}"
                ruta_tmp.write_bytes(archivo.read())

            doc = Documento(
                clave=key,
                nombre_archivo=f"{alumno.curp}_{nombre_archivo}",
                ruta_origen=ruta_tmp,
                presentado=bool(entregado),
            )
            alumno.docs[key] = doc

            pdf_path = procesar_documento_a_pdf(doc, carpeta_alumno)
            if pdf_path and pdf_path not in pdfs_alumno:
                pdfs_alumno.append(pdf_path)

        # Generar PDF completo del alumno
        expediente_pdf_path = None
        if pdfs_alumno:
            expediente_pdf_path = carpeta_alumno / f"{alumno.curp}_EXPEDIENTE_COMPLETO.pdf"
            if unir_pdfs_en_uno(pdfs_alumno, expediente_pdf_path):
                st.success(f"📄 Expediente completo generado: {expediente_pdf_path.name}")
            else:
                expediente_pdf_path = None
                st.warning("No se pudo crear el expediente completo del alumno (PDF).")

        # Actualizar Excels en memoria
        engine.agregar_alumno(alumno)
        machote_bytes, gestores_bytes = engine.get_excels_as_bytes()

        st.success("✅ Alumno procesado correctamente.")

        # Zona de descargas
        st.subheader("⬇️ Descargas")

        if expediente_pdf_path and expediente_pdf_path.exists():
            with open(expediente_pdf_path, "rb") as f:
                st.download_button(
                    "Descargar EXPEDIENTE COMPLETO (PDF)",
                    data=f.read(),
                    file_name=expediente_pdf_path.name,
                    mime="application/pdf",
                )

        st.download_button(
            "Descargar MACHOTE DE TRÁMITES (Excel)",
            data=machote_bytes.getvalue(),
            file_name="MACHOTE_DE_TRAMITES_AUTOMATICO.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.download_button(
            "Descargar FORMATO PARA LOS GESTORES (Excel)",
            data=gestores_bytes.getvalue(),
            file_name="FORMATO_PARA_LOS_GESTORES_AUTOMATICO.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.info("Puedes cambiar los datos y volver a presionar el botón para capturar otro alumno dentro de la misma sesión.")

st.markdown("---")
st.caption("Desarrollado por **AI Softwares® – Arquitectos de Ideas**")
